"""Agregacion de datos y generacion de reportes (HTML + Excel) con alertas."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from .reader import _to_number, detect_numeric_columns
from .alerts import evaluate_alerts


@dataclass
class Summary:
    title: str
    generated_at: str
    row_count: int
    group_by: str | None
    metrics: list[str]
    groups: list[dict] = field(default_factory=list)
    totals: dict = field(default_factory=dict)
    alerts: list[dict] = field(default_factory=list)


def summarize(
    headers: list[str],
    rows: list[dict],
    *,
    title: str = "Reporte de turno",
    group_by: str | None = None,
    metrics: list[str] | None = None,
    rules: list[str] | None = None,
) -> Summary:
    """Agrega filas por una columna, suma metricas y evalua reglas de alerta."""
    if metrics is None:
        auto = detect_numeric_columns(headers, rows)
        metrics = [m for m in auto if m != group_by]

    for col in filter(None, [group_by, *metrics]):
        if col not in headers:
            raise ValueError(f"La columna '{col}' no existe. Disponibles: {headers}")

    totals = {m: 0.0 for m in metrics}
    grouped: dict[str, dict] = {}

    for r in rows:
        for m in metrics:
            num = _to_number(r.get(m))
            if num is not None:
                totals[m] += num
        key = str(r.get(group_by)) if group_by else "Total"
        bucket = grouped.setdefault(key, {"key": key, "count": 0, "sums": {m: 0.0 for m in metrics}})
        bucket["count"] += 1
        for m in metrics:
            num = _to_number(r.get(m))
            if num is not None:
                bucket["sums"][m] += num

    groups = sorted(grouped.values(), key=lambda g: g["count"], reverse=True)
    alerts = evaluate_alerts(headers, rows, rules) if rules else []

    return Summary(
        title=title,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        row_count=len(rows),
        group_by=group_by,
        metrics=metrics,
        groups=groups,
        totals={"count": len(rows), "sums": totals},
        alerts=alerts,
    )


def _fmt(n: float) -> str:
    if n == int(n):
        return f"{int(n):,}".replace(",", ".")
    return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def build_html(summary: Summary) -> str:
    """Genera un reporte HTML autocontenido (tema claro, apto para email)."""
    metric_headers = "".join(f"<th>{m}</th>" for m in summary.metrics)

    rows_html = ""
    for g in summary.groups:
        cells = "".join(f"<td>{_fmt(g['sums'][m])}</td>" for m in summary.metrics)
        rows_html += f"<tr><td class='k'>{g['key']}</td><td>{_fmt(g['count'])}</td>{cells}</tr>"

    totals_cells = "".join(f"<td>{_fmt(summary.totals['sums'][m])}</td>" for m in summary.metrics)
    group_label = summary.group_by or "Grupo"

    n_alerts = len(summary.alerts)
    alert_card = ""
    if summary.alerts:
        alert_card = f'<div class="card alert"><div class="n">{n_alerts}</div><div class="l">Alertas</div></div>'

    cards = (
        f'<div class="card"><div class="n">{_fmt(summary.row_count)}</div><div class="l">Registros</div></div>'
        f'<div class="card"><div class="n">{len(summary.groups)}</div><div class="l">{group_label}s</div></div>'
        f'{alert_card}'
    )
    for m in summary.metrics[:2]:
        cards += f'<div class="card"><div class="n">{_fmt(summary.totals["sums"][m])}</div><div class="l">{m}</div></div>'

    alerts_block = ""
    if summary.alerts:
        arows = ""
        for a in summary.alerts:
            ctx = a["row"].get(summary.group_by) if summary.group_by else f"Fila {a['index'] + 1}"
            arows += (
                f"<tr><td class='k'>{ctx}</td><td>{a['metric']}</td>"
                f"<td class='bad'>{_fmt(a['value'])}</td><td>{a['op']} {_fmt(a['threshold'])}</td></tr>"
            )
        alerts_block = (
            '<div class="alerts"><h2>&#9888; Alertas (' + str(n_alerts) + ')</h2>'
            '<table><thead><tr><th>' + group_label + '</th><th>Metrica</th>'
            '<th>Valor</th><th>Regla</th></tr></thead><tbody>' + arows + '</tbody></table></div>'
        )

    style = """
  body{font-family:'Segoe UI',Arial,sans-serif;background:#f4f7fa;color:#1b2836;margin:0;padding:32px}
  .wrap{max-width:860px;margin:0 auto;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 6px 24px rgba(20,40,60,.08)}
  .head{background:#0f1720;color:#fff;padding:28px 36px}
  .head h1{margin:0;font-size:26px}
  .head p{margin:6px 0 0;color:#8aa0b2;font-size:14px}
  .cards{display:flex;gap:16px;padding:24px 36px;flex-wrap:wrap}
  .card{flex:1;min-width:130px;background:#f0f5f8;border:1px solid #e1e9ef;border-radius:12px;padding:18px}
  .card .n{font-size:30px;font-weight:800;color:#0ea5a0}
  .card .l{color:#5a6b7a;font-size:13px;margin-top:4px}
  .card.alert{background:#fdecec;border-color:#f5c2c2}
  .card.alert .n{color:#d64545}
  table{width:100%;border-collapse:collapse;margin:8px 0 0}
  th,td{padding:12px 36px;text-align:right;font-size:14px;border-bottom:1px solid #eef2f5}
  th:first-child,td:first-child{text-align:left}
  thead th{background:#f7fafc;color:#5a6b7a;text-transform:uppercase;font-size:12px;letter-spacing:.05em}
  td.k{font-weight:600}
  td.bad{color:#d64545;font-weight:800}
  tfoot td{font-weight:800;background:#0f1720;color:#fff}
  .alerts{padding:20px 0 0}
  .alerts h2{padding:20px 36px 0;margin:0;font-size:18px;color:#d64545}
  .foot{padding:18px 36px;color:#8aa0b2;font-size:12px;border-top:1px solid #eef2f5}
"""

    return (
        "<!DOCTYPE html>\n<html lang=\"es\"><head><meta charset=\"UTF-8\">"
        f"<title>{summary.title}</title>\n<style>{style}</style></head>\n"
        "<body><div class=\"wrap\">\n"
        f'  <div class="head"><h1>{summary.title}</h1><p>Generado el {summary.generated_at}</p></div>\n'
        f'  <div class="cards">{cards}</div>\n'
        "  <table>\n"
        f"    <thead><tr><th>{group_label}</th><th>Registros</th>{metric_headers}</tr></thead>\n"
        f"    <tbody>{rows_html}</tbody>\n"
        f'    <tfoot><tr><td>Total</td><td>{_fmt(summary.row_count)}</td>{totals_cells}</tr></tfoot>\n'
        "  </table>\n"
        f"  {alerts_block}\n"
        '  <div class="foot">ShiftReport Generator - Cesar Romero del Palacio</div>\n'
        "</div></body></html>"
    )


def write_html(summary: Summary, out_path: str | Path) -> Path:
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(build_html(summary), encoding="utf-8")
    return p


def write_excel(summary: Summary, rows: list[dict], headers: list[str], out_path: str | Path) -> Path:
    """Genera un Excel con hoja de resumen, detalle y (si hay) alertas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="0F1720")
    head_font = Font(color="FFFFFF", bold=True)
    accent_font = Font(color="0EA5A0", bold=True)
    bad_fill = PatternFill("solid", fgColor="FDECEC")
    bad_font = Font(color="D64545", bold=True)

    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = summary.title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Generado el {summary.generated_at}"
    ws["A2"].font = Font(color="8AA0B2")
    if summary.alerts:
        ws["A3"] = f"ALERTAS: {len(summary.alerts)}"
        ws["A3"].font = bad_font

    header = [summary.group_by or "Grupo", "Registros", *summary.metrics]
    ws.append([])
    ws.append(header)
    for cell in ws[ws.max_row]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center")

    for g in summary.groups:
        ws.append([g["key"], g["count"], *[g["sums"][m] for m in summary.metrics]])

    ws.append(["TOTAL", summary.row_count, *[summary.totals["sums"][m] for m in summary.metrics]])
    for cell in ws[ws.max_row]:
        cell.font = accent_font

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 4, 12), 40)

    wd = wb.create_sheet("Detalle")
    wd.append(headers)
    for cell in wd[1]:
        cell.fill = head_fill
        cell.font = head_font
    for r in rows:
        wd.append([r.get(h) for h in headers])
    wd.freeze_panes = "A2"

    if summary.alerts:
        wa = wb.create_sheet("Alertas")
        wa.append(["Regla", "Metrica", "Valor", *headers])
        for cell in wa[1]:
            cell.fill = head_fill
            cell.font = head_font
        for a in summary.alerts:
            wa.append([a["rule"], a["metric"], a["value"], *[a["row"].get(h) for h in headers]])
            for cell in wa[wa.max_row]:
                cell.fill = bad_fill
        wa.freeze_panes = "A2"

    wb.save(p)
    return p
